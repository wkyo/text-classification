# coding: utf-8
import os
import csv

import numpy as np
from openpyxl import load_workbook
from keras_preprocessing.sequence import pad_sequences
from sklearn.model_selection import train_test_split


def load_categorical_text_from_csv(path,
                                   fields=None,
                                   field_label='label',
                                   field_text='text'):
    """Load all categorical text from csv"""
    texts, labels = [], []
    with open(path, 'rt', encoding='utf-8') as fp:
        reader = csv.DictReader(fp, fieldnames=fields)
        for d in reader:
            texts.append(d[field_text])
            labels.append(d[field_label])
    return texts, labels


def load_categorical_text_from_excel(path,
                                     fields=None,
                                     field_label='label',
                                     field_text='text'):
    """Load categorical text from excel

    data type is train, validation and test
    """
    texts, labels = [], []
    workbook = load_workbook(path)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    if fields is None:
        fields = [str(cell) for cell in next(row_iter)]
    for row in row_iter:
        d = {k: v for k, v in zip(fields, row)}
        texts.append(d[field_text])
        labels.append(d[field_label])
    return texts, labels


def normalize_texts(*texts_array, tokenizer, tokenizer_update=True, maxlen=None):
    if tokenizer_update:
        for texts in texts_array:
            tokenizer.fit_on_texts(texts)

    sequences_array = []
    for texts in texts_array:
        if texts:
            sequences = tokenizer.texts_to_sequences(texts)
            sequences = pad_sequences(sequences, maxlen=maxlen)
            sequences_array.append(sequences)
        else:
            sequences_array.append(None)

    return sequences_array


def unique_labels(*labels_array):
    labels = set()
    for l in labels_array:
        labels = labels.union(l)
    labels = sorted(labels)
    return labels


def load_dataset(path,
                 tokenizer=None,
                 tokenizer_update=True,
                 test_size=0.2,
                 labels=None,
                 maxlen=None,
                 field_label='label',
                 field_text='text'):
    if os.path.isfile(path):
        suffix = os.path.splitext(path)[-1].lower()
        if suffix == '.csv':
            data = load_categorical_text_from_csv(
                path,
                field_label=field_label,
                field_text=field_text
            )
        elif suffix in ('.xls', '.xlsx'):
            data = load_categorical_text_from_excel(
                path,
                field_label=field_label,
                field_text=field_text
            )
        else:
            raise ValueError('Unsupported file type')
        x_trn, x_tst, y_trn, y_tst = train_test_split(
            *data, test_size=test_size)
        train_data = (x_trn, y_trn)
        validation_data = [], []
        test_data = (x_tst, y_tst)
    else:
        train_data = load_categorical_text_from_csv(
            os.path.join(path, 'train.csv'),
            field_label=field_label,
            field_text=field_text
        )
        try:
            validation_data = load_categorical_text_from_csv(
                os.path.join(path, 'validate.csv'),
                field_label=field_label,
                field_text=field_text
            )
        except IOError:
            validation_data = [], []
        test_data = load_categorical_text_from_csv(
            os.path.join(path, 'test.csv'))

    (x_trn, y_trn), (x_val, y_val), (x_tst,
                                     y_tst) = train_data, validation_data, test_data

    # pylint: disable=unbalanced-tuple-unpacking
    x_trn, x_val, x_tst = normalize_texts(
        x_trn, x_val, x_tst,
        tokenizer=tokenizer,
        tokenizer_update=tokenizer_update,
        maxlen=maxlen
    )

    if not labels:
        labels = unique_labels(y_trn, y_val, y_tst)
    y_trn = np.array([labels.index(y) for y in y_trn])
    y_val = np.array([labels.index(y) for y in y_val])
    y_tst = np.array([labels.index(y) for y in y_tst])

    if not x_val:
        validation_data = None
    else:
        validation_data = (x_val, y_val)

    data = ((x_trn, y_trn), validation_data, (x_tst, y_tst))
    data_dict = {
        'classes': labels,
        'word_index': tokenizer.word_index
    }

    return data, data_dict
